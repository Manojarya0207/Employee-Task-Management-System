import io
import csv
from datetime import date, datetime
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from models.task import Task
from models.employee import Employee
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from repositories.task_repository import TaskRepository
from datetime import timedelta

def fetch_report_data(
    db: Session, 
    report_type: str, 
    start_date: date = None, 
    end_date: date = None, 
    employee_id: str = None, 
    status: str = None
) -> list[dict]:
    """
    Fetches task records matching report criteria joined with Employee details.
    """
    today = date.today()
    if report_type == 'daily':
        start_date = today
        end_date = today
    elif report_type == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
    elif report_type == 'monthly':
        start_date = today - timedelta(days=30)
        end_date = today

    results = TaskRepository.get_tasks_joined_with_employees(
        db, start_date=start_date, end_date=end_date, employee_id=employee_id, status=status
    )
    
    data = []
    for task, emp in results:
        data.append({
            'task_id': task.task_id,
            'date': task.created_date.strftime('%Y-%m-%d'),
            'time': task.created_time.strftime('%I:%M %p'),
            'employee_id': emp.employee_id,
            'employee_name': emp.employee_name,
            'department': emp.department or 'N/A',
            'title': task.title,
            'description': task.description or '',
            'status': task.status,
            'last_modified': task.last_modified.strftime('%Y-%m-%d %I:%M %p')
        })
    return data


def export_csv(data: list[dict]) -> io.BytesIO:
    """Generates a CSV file buffer from report data."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        'Task ID', 'Date', 'Time', 'Employee ID', 
        'Employee Name', 'Department', 'Task Title', 
        'Description', 'Status', 'Last Modified'
    ])
    
    for row in data:
        writer.writerow([
            row['task_id'], row['date'], row['time'], row['employee_id'],
            row['employee_name'], row['department'], row['title'],
            row['description'], row['status'], row['last_modified']
        ])
        
    buffer = io.BytesIO()
    buffer.write(output.getvalue().encode('utf-8'))
    buffer.seek(0)
    return buffer

def export_excel(data: list[dict], title: str) -> io.BytesIO:
    """Generates a beautifully styled Excel workbook using openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TaskFlow Report"
    
    # Theme colors
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo
    zebra_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid") # Light Violet
    border_color = Side(border_style="thin", color="E5E7EB")
    border = Border(left=border_color, right=border_color, top=border_color, bottom=border_color)
    
    # Report Title Block
    ws.merge_cells("A1:J1")
    ws["A1"] = f"TaskFlow Management Report - {title}"
    ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="4F46E5")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"
    ws["A2"].font = Font(name="Segoe UI", size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20
    
    # Headers
    headers = [
        'Task ID', 'Date', 'Time', 'Employee ID', 
        'Employee Name', 'Department', 'Task Title', 
        'Description', 'Status', 'Last Modified'
    ]
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = text
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        
    ws.row_dimensions[4].height = 28
    
    # Data rows
    for row_idx, item in enumerate(data, 5):
        row_values = [
            item['task_id'], item['date'], item['time'], item['employee_id'],
            item['employee_name'], item['department'], item['title'],
            item['description'], item['status'], item['last_modified']
        ]
        
        is_even = row_idx % 2 == 0
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = border
            
            # Alignments
            if col_idx in [1, 2, 3, 4, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            # Zebra striping
            if is_even:
                cell.fill = zebra_fill
                
        ws.row_dimensions[row_idx].height = 22
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Skip title row in width calculation
            if cell.row in [1, 2]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_pdf(data: list[dict], title: str) -> io.BytesIO:
    """Generates a landscape PDF document containing a styled report table."""
    buffer = io.BytesIO()
    
    # We use landscape to fit all the columns comfortably
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=6
    )
    
    meta_style = ParagraphStyle(
        'DocMeta',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=9,
        textColor=colors.HexColor('#6B7280'),
        spaceAfter=15
    )
    
    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10
    )
    
    cell_bold_style = ParagraphStyle(
        'TableCellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10
    )
    
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white
    )

    story = []
    
    # Title & Metadata
    story.append(Paragraph(f"TaskFlow Management Report - {title}", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}", meta_style))
    story.append(Spacer(1, 10))
    
    # Table headers
    headers = [
        Paragraph('ID', header_style),
        Paragraph('Date', header_style),
        Paragraph('Employee ID', header_style),
        Paragraph('Employee Name', header_style),
        Paragraph('Dept', header_style),
        Paragraph('Task Title', header_style),
        Paragraph('Description', header_style),
        Paragraph('Status', header_style),
        Paragraph('Modified', header_style)
    ]
    
    table_data = [headers]
    
    # Data rows
    for item in data:
        row = [
            Paragraph(str(item['task_id']), cell_style),
            Paragraph(f"{item['date']}<br/>{item['time']}", cell_style),
            Paragraph(item['employee_id'], cell_style),
            Paragraph(item['employee_name'], cell_bold_style),
            Paragraph(item['department'], cell_style),
            Paragraph(item['title'], cell_bold_style),
            Paragraph(item['description'][:150] + ('...' if len(item['description']) > 150 else ''), cell_style),
            Paragraph(item['status'], cell_style),
            Paragraph(item['last_modified'], cell_style)
        ]
        table_data.append(row)
        
    # Table Widths (landscape letter width is 792. Margins are 36 each side, so printable area is 720)
    col_widths = [30, 65, 60, 80, 55, 120, 190, 60, 60]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Build Table Style
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
    ])
    
    # Apply zebra striping
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F3FF'))
            
    t.setStyle(t_style)
    story.append(t)
    
    doc.build(story)
    buffer.seek(0)
    return buffer

# Helper timedelta needed for date math
from datetime import timedelta
